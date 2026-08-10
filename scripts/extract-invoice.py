#!/usr/bin/env python3
"""請求書ファイル1件から表を抽出し、admin_invoice_stage_import に渡せるJSONを標準出力に出す。

対応形式:
  .xlsx           openpyxl（表がそのまま取れる。最も確実）
  .pdf(テキスト)   pdftotext -layout の出力から列を推定
  画像・画像PDF    対象外（exit 3）。画像認識で読み取ってJSONを手で用意する

抽出は「候補を出す」ことが仕事で、正しさの確定は人が確認画面で行う。
判別に自信が無い値は入れず、source_ref に元の場所（シート/行・ページ/行）を必ず残す。
"""
import sys, os, re, json, subprocess, datetime

ITEM_HEADERS = ('品名', '商品名', '品目', '内容')
QTY_HEADERS = ('数量', '重量', 'kg', 'キロ', '数')
PRICE_HEADERS = ('単価',)
AMOUNT_HEADERS = ('金額', '合計', '価格')

def norm(v):
    if v is None: return ''
    return str(v).strip()

def to_num(v):
    s = re.sub(r'[,¥円\s]', '', norm(v))
    if not s: return None
    m = re.match(r'^-?\d+(\.\d+)?$', s)
    return float(s) if m else None

def to_date(v):
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.strftime('%Y-%m-%d')
    s = norm(v)
    m = re.search(r'(\d{4})[/年.-](\d{1,2})[/月.-](\d{1,2})', s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2000 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31:
            return f'{y}-{mo:02d}-{d:02d}'
    m = re.search(r'令和\s*(\d+)\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', s)
    if m:
        mo, d = int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f'{2018+int(m.group(1))}-{mo:02d}-{d:02d}'
    return None

def find_phone(text):
    # 「TEL 0287-39-5529」形式。空白は区切りとみなし電話の中には含めない
    t = text.translate(str.maketrans('０１２３４５６７８９', '0123456789'))
    m = re.search(r'(?:TEL|Tel|tel|電話)[:：\s]*(0[\d\-‐()（）]{8,13})', t)
    if not m:
        m = re.search(r'(?<![\d-])(0\d{1,4}-\d{1,4}-\d{3,4})(?![\d-])', t)
    if m:
        digits = re.sub(r'\D', '', m.group(1))
        if 10 <= len(digits) <= 11: return m.group(1).strip()
    return None

def find_postal(text):
    m = re.search(r'〒\s*(\d{3}[-‐]?\d{4})', text)
    return m.group(1) if m else None

def extract_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    docs = []
    for sheet in wb.worksheets:
        rows = [[c.value for c in r] for r in sheet.iter_rows(max_row=min(sheet.max_row, 500))]
        header_idx, cols = None, {}
        for i, row in enumerate(rows):
            texts = [norm(v) for v in row]
            if any(any(h in t for h in ITEM_HEADERS) for t in texts) and \
               any(any(h in t for h in PRICE_HEADERS + AMOUNT_HEADERS) for t in texts):
                for j, t in enumerate(texts):
                    if not t: continue
                    if any(h in t for h in ITEM_HEADERS): cols.setdefault('item', j)
                    elif any(h in t for h in QTY_HEADERS): cols.setdefault('qty', j)
                    elif any(h in t for h in PRICE_HEADERS): cols.setdefault('price', j)
                    elif any(h in t for h in AMOUNT_HEADERS): cols.setdefault('amount', j)
                header_idx = i
                break
        if header_idx is None or 'item' not in cols:
            continue
        head_text = '\n'.join(' '.join(norm(v) for v in row if norm(v)) for row in rows[:header_idx])
        doc = {
            'raw_customer_name': None, 'raw_addressee': None,
            'raw_phone': find_phone(head_text), 'raw_postal': find_postal(head_text),
            'invoice_number': None, 'invoice_date': None,
            'note': f'sheet:{sheet.title}', 'lines': [],
        }
        for row in rows[:header_idx]:
            for v in row:
                t = norm(v)
                if re.search(r'(様|御中)\s*$', t) and not doc['raw_addressee']:
                    doc['raw_addressee'] = t
                    doc['raw_customer_name'] = re.sub(r'\s*(様|御中)\s*$', '', t)
                m = re.match(r'^(?:No\.?|請求書番号)[:：\s]*(\S+)$', t)
                if m and not doc['invoice_number']: doc['invoice_number'] = m.group(1)
                d = to_date(v)
                if d and not doc['invoice_date']: doc['invoice_date'] = d
        total = None
        for i in range(header_idx + 1, len(rows)):
            row = rows[i]
            item = norm(row[cols['item']]) if cols['item'] < len(row) else ''
            if not item: continue
            if re.match(r'^(小計|合計|消費税|税込|総計)', item):
                a = to_num(row[cols.get('amount', cols.get('price', 0))]) if row else None
                if item.startswith(('合計', '総計')) and a: total = a
                continue
            line = {'raw_item_name': item, 'source_ref': f'sheet:{sheet.title} 行{i+1}'}
            q = to_num(row[cols['qty']]) if 'qty' in cols and cols['qty'] < len(row) else None
            if q is not None: line['weight_kg'] = q
            p = to_num(row[cols['price']]) if 'price' in cols and cols['price'] < len(row) else None
            if p is not None: line['unit_price'] = p
            a = to_num(row[cols['amount']]) if 'amount' in cols and cols['amount'] < len(row) else None
            if a is not None: line['amount'] = a
            g = re.search(r'[（(](並|上|極上)[）)]', item)
            if g: line['raw_grade'] = g.group(1)
            doc['lines'].append(line)
        if total is not None: doc['total_amount'] = total
        if doc['lines']: docs.append(doc)
    return docs, len(wb.worksheets)

def extract_pdf(path):
    out = subprocess.run(['pdftotext', '-layout', path, '-'], capture_output=True, text=True)
    if out.returncode != 0: raise RuntimeError(out.stderr[:200])
    text = out.stdout
    if len(re.sub(r'\s', '', text)) < 40:
        # ほぼ文字が取れない = 画像PDF。画像認識の経路へ回す
        sys.stderr.write('image-pdf\n'); sys.exit(3)
    pages = text.split('\f')
    docs = []
    for pi, page in enumerate(pages, start=1):
        lines = [l for l in page.splitlines() if l.strip()]
        if not lines: continue
        head = '\n'.join(lines[:12])
        doc = {'page_from': pi, 'page_to': pi,
               'raw_phone': find_phone(head), 'raw_postal': find_postal(head),
               'invoice_date': to_date(head), 'invoice_number': None,
               'raw_customer_name': None, 'raw_addressee': None, 'lines': []}
        for l in lines[:12]:
            m = re.search(r'(\S[^\s]{1,40}?)\s*(様|御中)\s*$', l.strip())
            if m and not doc['raw_addressee']:
                doc['raw_addressee'] = m.group(0).strip()
                doc['raw_customer_name'] = m.group(1)
            m = re.search(r'(?:No\.?|請求書番号)[:：\s]*([A-Za-z0-9\-]+)', l)
            if m and not doc['invoice_number']: doc['invoice_number'] = m.group(1)
        for li, l in enumerate(lines, start=1):
            # 「品名 …… 数量 単価 金額」のならびを右から数値3つで推定
            nums = re.findall(r'[\d,]+(?:\.\d+)?', l.replace('¥', ''))
            name = re.split(r'\s{2,}', l.strip())[0]
            if len(nums) >= 2 and name and not re.match(r'^(小計|合計|消費税|税込|お振込|〒|TEL|電話)', name) \
               and not re.match(r'^[\d,._%¥ -]+$', name):
                vals = [to_num(n) for n in nums[-3:]]
                line = {'raw_item_name': name, 'source_ref': f'p.{pi} 行{li}', 'confidence': 0.6}
                if len(vals) == 3: line['weight_kg'], line['unit_price'], line['amount'] = vals
                else: line['unit_price'], line['amount'] = vals[-2], vals[-1]
                doc['lines'].append(line)
        if doc['lines']: docs.append(doc)
    return docs, len(pages)

def main():
    path = sys.argv[1]
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        docs, pages = extract_xlsx(path)
    elif ext == '.pdf':
        docs, pages = extract_pdf(path)
    elif ext in ('.png', '.jpg', '.jpeg', '.heic'):
        sys.stderr.write('image\n'); sys.exit(3)
    else:
        sys.stderr.write(f'unsupported:{ext}\n'); sys.exit(2)
    print(json.dumps({'documents': docs, 'page_count': pages}, ensure_ascii=False))

if __name__ == '__main__':
    main()
